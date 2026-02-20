#!/usr/bin/env python3
"""
Bulk Indexing Script for Dropbox Files
======================================
Downloads files from Dropbox, extracts text, generates embeddings,
and stores everything in your database. Resumes from where it left off.

Usage:
  1. pip install openai supabase dropbox PyPDF2 python-docx openpyxl
  2. Set environment variables (or edit the values below):
     - SUPABASE_URL
     - SUPABASE_SERVICE_ROLE_KEY
     - OPENAI_API_KEY
     - DROPBOX_ACCESS_TOKEN
  3. python bulk_index.py
"""

import os
import re
import io
import sys
import csv
import time
import email
import traceback
from pathlib import Path

# ---------------------------------------------------------------------------
# Configuration – set these as env vars or paste values directly
# ---------------------------------------------------------------------------
SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
DROPBOX_ACCESS_TOKEN = os.getenv("DROPBOX_ACCESS_TOKEN", "")

CHUNK_SIZE = 1000
CHUNK_OVERLAP = 200
EMBEDDING_BATCH_SIZE = 5
DELAY_BETWEEN_FILES = 0.5  # seconds

# Extensions that CANNOT be vectorized – will be marked "skipped"
SKIP_EXTENSIONS = {
    # Images
    "jpg", "jpeg", "png", "gif", "bmp", "tif", "tiff", "heic", "heif",
    "dng", "raw", "cr2", "nef", "svg", "ico", "webp",
    # Video
    "mov", "mp4", "avi", "wmv", "mkv", "flv", "m4v",
    # Audio
    "mp3", "wav", "aac", "flac", "ogg", "wma",
    # CAD / GIS
    "dwg", "dgn", "dxf", "shp", "shx", "dbf", "kmz", "kml",
    # Archives
    "zip", "rar", "7z", "tar", "gz", "bz2",
    # Fonts
    "ttf", "otf", "woff", "woff2",
    # Other binary
    "bak", "mjs", "exe", "dll", "iso", "bin", "dat",
}

# ---------------------------------------------------------------------------
# Regex patterns for metadata extraction (mirrors process-document function)
# ---------------------------------------------------------------------------
COST_PATTERN = re.compile(r'\$[\d,]+(?:\.\d{2})?')
DATE_PATTERN = re.compile(
    r'\b(?:0?[1-9]|1[0-2])[-/](?:0?[1-9]|[12]\d|3[01])[-/](?:19|20)?\d{2}\b'
    r'|'
    r'\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},?\s+\d{4}\b',
    re.IGNORECASE
)
PROJECT_PATTERN = re.compile(r'(?:project|lot|tract|phase|unit|parcel)[\s:#-]*([A-Za-z0-9-]+)', re.IGNORECASE)
PERMIT_PATTERN = re.compile(r'(?:permit|license|bond)[\s:#-]*([A-Za-z0-9-]+)', re.IGNORECASE)
EXPIRY_PATTERN = re.compile(r'(?:expir(?:es?|ation|y)|due|valid until|renew(?:al)?)[:\s]+([^\n,;]+)', re.IGNORECASE)


# ===================================================================
# Text extraction helpers
# ===================================================================

def extract_pdf(data: bytes) -> str:
    from PyPDF2 import PdfReader
    reader = PdfReader(io.BytesIO(data))
    parts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            parts.append(text)
    return "\n".join(parts)


def extract_docx(data: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs)


def extract_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = "\t".join(cells).strip()
            if line:
                parts.append(line)
    wb.close()
    return "\n".join(parts)


def extract_csv_text(data: bytes) -> str:
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    parts = []
    for row in reader:
        line = "\t".join(row).strip()
        if line:
            parts.append(line)
    return "\n".join(parts)


def extract_eml(data: bytes) -> str:
    msg = email.message_from_bytes(data)
    parts = []
    subject = msg.get("Subject", "")
    if subject:
        parts.append(f"Subject: {subject}")
    sender = msg.get("From", "")
    if sender:
        parts.append(f"From: {sender}")
    date_str = msg.get("Date", "")
    if date_str:
        parts.append(f"Date: {date_str}")
    # Body
    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            if ct == "text/plain":
                payload = part.get_payload(decode=True)
                if payload:
                    parts.append(payload.decode("utf-8", errors="replace"))
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            parts.append(payload.decode("utf-8", errors="replace"))
    return "\n".join(parts)


def extract_plain_text(data: bytes) -> str:
    return data.decode("utf-8", errors="replace")


def extract_text(data: bytes, ext: str) -> str:
    """Route to the correct extractor based on file extension."""
    ext = ext.lower().lstrip(".")
    if ext == "pdf":
        return extract_pdf(data)
    elif ext == "docx":
        return extract_docx(data)
    elif ext in ("xlsx", "xls"):
        return extract_xlsx(data)
    elif ext == "csv":
        return extract_csv_text(data)
    elif ext in ("eml", "msg"):
        return extract_eml(data)
    elif ext in ("txt", "log", "md", "json", "xml", "html", "htm", "rtf"):
        return extract_plain_text(data)
    elif ext in ("doc", "pptx", "ppt"):
        # Attempt plain-text extraction as fallback
        return extract_plain_text(data)
    else:
        return extract_plain_text(data)


# ===================================================================
# Metadata extraction (mirrors process-document edge function)
# ===================================================================

def extract_metadata(text: str) -> dict:
    metadata = {}

    costs = COST_PATTERN.findall(text)
    if costs:
        nums = []
        for c in costs:
            try:
                nums.append(float(c.replace("$", "").replace(",", "")))
            except ValueError:
                pass
        if nums:
            metadata["costs"] = nums
            metadata["total_cost"] = sum(nums)
            metadata["max_cost"] = max(nums)
            metadata["min_cost"] = min(nums)

    dates = DATE_PATTERN.findall(text)
    if dates:
        metadata["dates"] = list(set(dates))

    projects = [m.group(1) for m in PROJECT_PATTERN.finditer(text)]
    if projects:
        metadata["projects"] = list(set(projects))
        metadata["project_name"] = projects[0]

    permits = [m.group(1) for m in PERMIT_PATTERN.finditer(text)]
    if permits:
        metadata["permits"] = list(set(permits))

    expirations = [m.group(1).strip() for m in EXPIRY_PATTERN.finditer(text)]
    if expirations:
        metadata["expirations"] = list(set(expirations))

    lower = text.lower()
    if "invoice" in lower or "billing" in lower:
        metadata["doc_type"] = "invoice"
    elif "permit" in lower or "license" in lower:
        metadata["doc_type"] = "permit"
    elif "contract" in lower or "agreement" in lower:
        metadata["doc_type"] = "contract"
    elif "proposal" in lower or "quote" in lower:
        metadata["doc_type"] = "proposal"
    elif "report" in lower or "summary" in lower:
        metadata["doc_type"] = "report"

    return metadata


# ===================================================================
# Chunking (mirrors process-document edge function)
# ===================================================================

def split_text(text: str) -> list[str]:
    separators = ["\n\n", "\n", ". ", " ", ""]

    def split_recursive(txt: str, sep_idx: int) -> list[str]:
        if len(txt) <= CHUNK_SIZE:
            return [txt]
        sep = separators[sep_idx]
        parts = txt.split(sep) if sep else list(txt)
        result = []
        current = ""
        for part in parts:
            part_with_sep = part + sep if sep else part
            if len(current) + len(part_with_sep) <= CHUNK_SIZE:
                current += part_with_sep
            else:
                if current:
                    result.append(current.strip())
                if len(part_with_sep) > CHUNK_SIZE and sep_idx < len(separators) - 1:
                    result.extend(split_recursive(part_with_sep, sep_idx + 1))
                    current = ""
                else:
                    current = part_with_sep
        if current.strip():
            result.append(current.strip())
        return result

    raw = split_recursive(text, 0)
    chunks = []
    for i, chunk in enumerate(raw):
        if i > 0 and CHUNK_OVERLAP > 0:
            overlap = raw[i - 1][-CHUNK_OVERLAP:]
            chunks.append(overlap + chunk)
        else:
            chunks.append(chunk)
    return [c for c in chunks if c.strip()]


# ===================================================================
# OpenAI embedding with retry
# ===================================================================

def generate_embedding(text: str, client) -> list[float]:
    for attempt in range(4):
        try:
            resp = client.embeddings.create(model="text-embedding-3-small", input=text)
            return resp.data[0].embedding
        except Exception as e:
            err_str = str(e)
            if ("429" in err_str or "500" in err_str or "502" in err_str or "503" in err_str) and attempt < 3:
                delay = 2 ** attempt
                print(f"  ⚠ OpenAI rate-limit/error, retrying in {delay}s…")
                time.sleep(delay)
            else:
                raise
    raise RuntimeError("Max retries exceeded for embedding")


def generate_embeddings_batch(texts: list[str], client) -> list[list[float]]:
    results = []
    for i in range(0, len(texts), EMBEDDING_BATCH_SIZE):
        batch = texts[i : i + EMBEDDING_BATCH_SIZE]
        for t in batch:
            results.append(generate_embedding(t, client))
        if i + EMBEDDING_BATCH_SIZE < len(texts):
            time.sleep(0.2)
    return results


# ===================================================================
# Database helpers
# ===================================================================

def update_indexing_status(supabase, file_path, file_name, status, chunks_created, error_message, metadata):
    from datetime import datetime, timezone
    record = {
        "file_path": file_path,
        "file_name": file_name,
        "status": status,
        "chunks_created": chunks_created,
        "error_message": error_message,
        "metadata": metadata,
        "indexed_at": datetime.now(timezone.utc).isoformat() if status == "success" else None,
    }
    supabase.table("indexing_status").upsert(record, on_conflict="file_path").execute()


def fetch_unindexed_files(supabase) -> list[dict]:
    """Fetch ALL unindexed files via the RPC, paginating in batches of 1000."""
    all_files = []
    page_size = 1000
    offset = 0
    while True:
        resp = supabase.rpc("get_unindexed_dropbox_files", {
            "p_limit": page_size,
            "p_offset": offset,
        }).execute()
        batch = resp.data or []
        all_files.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
        print(f"  Fetched {len(all_files)} file records so far…")
    return all_files


# ===================================================================
# Main
# ===================================================================

def main():
    # Validate config
    missing = []
    if not SUPABASE_URL:
        missing.append("SUPABASE_URL")
    if not SUPABASE_SERVICE_KEY:
        missing.append("SUPABASE_SERVICE_ROLE_KEY")
    if not OPENAI_API_KEY:
        missing.append("OPENAI_API_KEY")
    if not DROPBOX_ACCESS_TOKEN:
        missing.append("DROPBOX_ACCESS_TOKEN")
    if missing:
        print(f"❌ Missing required config: {', '.join(missing)}")
        print("Set them as environment variables or edit the values at the top of this script.")
        sys.exit(1)

    # Initialize clients
    from supabase import create_client
    from openai import OpenAI
    import dropbox

    supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)
    openai_client = OpenAI(api_key=OPENAI_API_KEY)
    dbx = dropbox.Dropbox(DROPBOX_ACCESS_TOKEN)

    # Test Dropbox connection
    try:
        account = dbx.users_get_current_account()
        print(f"✅ Connected to Dropbox as: {account.name.display_name}")
    except Exception as e:
        print(f"❌ Dropbox connection failed: {e}")
        sys.exit(1)

    # Fetch unindexed files
    print("\n📂 Fetching unindexed files from database…")
    files = fetch_unindexed_files(supabase)
    print(f"   Found {len(files)} unindexed files\n")

    if not files:
        print("🎉 All files are already indexed! Nothing to do.")
        return

    # Separate into vectorizable vs skip
    to_process = []
    to_skip = []
    for f in files:
        ext = (f.get("file_extension") or "").lower().lstrip(".")
        if ext in SKIP_EXTENSIONS:
            to_skip.append(f)
        else:
            to_process.append(f)

    print(f"📊 Will process: {len(to_process)} files")
    print(f"⏭  Will skip:    {len(to_skip)} non-vectorizable files\n")

    # Mark non-vectorizable files as skipped
    if to_skip:
        print("Marking non-vectorizable files as skipped…")
        for i, f in enumerate(to_skip):
            fp = f["file_path"]
            fn = f.get("file_name")
            ext = (f.get("file_extension") or "").lower().lstrip(".")
            update_indexing_status(supabase, fp, fn, "skipped", 0, f"Non-vectorizable extension: {ext}", {})
            if (i + 1) % 200 == 0:
                print(f"  Skipped {i + 1}/{len(to_skip)}")
        print(f"  ✅ Marked {len(to_skip)} files as skipped\n")

    # Process vectorizable files
    success_count = 0
    fail_count = 0
    total = len(to_process)

    for idx, f in enumerate(to_process):
        fp = f["file_path"]
        fn = f.get("file_name") or Path(fp).name
        ext = (f.get("file_extension") or "").lower().lstrip(".")
        progress = f"[{idx + 1}/{total}]"

        try:
            # Download from Dropbox
            try:
                _, response = dbx.files_download(fp)
                file_data = response.content
            except Exception as e:
                err = f"Dropbox download failed: {e}"
                print(f"  {progress} ❌ {fn} — {err}")
                update_indexing_status(supabase, fp, fn, "failed", 0, err, {})
                fail_count += 1
                continue

            # Extract text
            try:
                content = extract_text(file_data, ext)
            except Exception as e:
                err = f"Text extraction failed: {e}"
                print(f"  {progress} ❌ {fn} — {err}")
                update_indexing_status(supabase, fp, fn, "failed", 0, err, {})
                fail_count += 1
                continue

            # Skip if too short
            if not content or len(content.strip()) < 50:
                print(f"  {progress} ⏭  {fn} — insufficient content ({len(content.strip()) if content else 0} chars)")
                update_indexing_status(supabase, fp, fn, "skipped", 0, "Insufficient content (< 50 chars)", {})
                continue

            # Chunk
            chunks = split_text(content)

            # Extract metadata
            extracted_meta = extract_metadata(content)

            # Generate embeddings
            embeddings = generate_embeddings_batch(chunks, openai_client)

            # Delete existing chunks for this file path
            supabase.table("documents").delete().eq("file_path", fp).execute()

            # Insert new chunks
            documents = []
            for i, (chunk, emb) in enumerate(zip(chunks, embeddings)):
                documents.append({
                    "content": chunk,
                    "embedding": str(emb),
                    "file_path": fp,
                    "file_name": fn,
                    "metadata": {
                        **extracted_meta,
                        "chunk_index": i,
                        "total_chunks": len(chunks),
                    },
                })

            # Insert in batches of 50 to avoid payload limits
            for b in range(0, len(documents), 50):
                batch = documents[b : b + 50]
                supabase.table("documents").insert(batch).execute()

            # Update status
            update_indexing_status(supabase, fp, fn, "success", len(chunks), None, extracted_meta)
            success_count += 1
            print(f"  {progress} ✅ {fn} ({len(chunks)} chunks)")

        except Exception as e:
            err = str(e)
            print(f"  {progress} ❌ {fn} — {err}")
            traceback.print_exc()
            update_indexing_status(supabase, fp, fn, "failed", 0, err, {})
            fail_count += 1

        # Delay between files
        time.sleep(DELAY_BETWEEN_FILES)

    print(f"\n{'='*50}")
    print(f"✅ Done! Processed {total} files")
    print(f"   Success: {success_count}")
    print(f"   Failed:  {fail_count}")
    print(f"   Skipped: {total - success_count - fail_count}")


if __name__ == "__main__":
    main()
